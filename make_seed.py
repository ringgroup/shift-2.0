from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date

ARIAL = "Arial"
BLUE = "0000FF"   # inputs
BLACK = "000000"  # formulas
AMBER = "1F2330"
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
TITLE_FILL = PatternFill("solid", fgColor="111827")
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ---------- Measurements ----------
ws = wb.active
ws.title = "Measurements"

ws["A1"] = "PROMETHEUS TRACER — Seed Data (Retatrutide protocol)"
ws["A1"].font = Font(name=ARIAL, bold=True, size=14, color="FFFFFF")
ws["A1"].fill = TITLE_FILL
ws.merge_cells("A1:H1")
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 24

ws["A3"] = "Start (Day 0)"
ws["B3"] = date(2026, 4, 14)
ws["A4"] = "Goal loss (kg)"
ws["B4"] = 40
for c in ("A3", "A4"):
    ws[c].font = Font(name=ARIAL, bold=True)
ws["B3"].font = Font(name=ARIAL, color=BLUE)
ws["B3"].number_format = "yyyy-mm-dd"
ws["B4"].font = Font(name=ARIAL, color=BLUE)

headers = ["Date", "Day", "Weight (kg)", "Muscle (kg)", "Body Fat (%)",
           "Fat Mass (kg)", "Total Lost (kg)", "% of Goal"]
hrow = 6
for i, h in enumerate(headers, start=1):
    cell = ws.cell(row=hrow, column=i, value=h)
    cell.font = Font(name=ARIAL, bold=True, color="FFFFFF")
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")
    cell.border = border

rows = [
    (date(2026, 4, 21), 154.7, 49.6, 44.4),
    (date(2026, 5, 2),  151.6, 50.7, 41.9),
    (date(2026, 5, 17), 150.3, 50.2, 42.0),
    (date(2026, 5, 26), 146.5, 50.9, 39.7),
    (date(2026, 6, 8),  142.6, 49.5, 39.6),
]

first = hrow + 1
start = date(2026, 4, 14)
baseline = rows[0][1]
goal = 40
for idx, (d, w, m, f) in enumerate(rows):
    r = first + idx
    ws.cell(row=r, column=1, value=d).number_format = "yyyy-mm-dd"
    ws.cell(row=r, column=2, value=(d - start).days)              # Day
    ws.cell(row=r, column=3, value=w)                              # Weight
    ws.cell(row=r, column=4, value=m)                              # Muscle
    ws.cell(row=r, column=5, value=f)                              # Body Fat %
    ws.cell(row=r, column=6, value=round(w * f / 100, 1))         # Fat Mass
    ws.cell(row=r, column=7, value=round(baseline - w, 1))        # Total Lost
    ws.cell(row=r, column=8, value=(baseline - w) / goal)         # % of goal
    # styling
    for col in range(1, 9):
        cc = ws.cell(row=r, column=col)
        cc.border = border
        cc.alignment = Alignment(horizontal="center")
        cc.font = Font(name=ARIAL, color=(BLUE if col in (3, 4, 5) else BLACK))
    ws.cell(row=r, column=3).number_format = "0.0"
    ws.cell(row=r, column=4).number_format = "0.0"
    ws.cell(row=r, column=5).number_format = '0.0"%"'
    ws.cell(row=r, column=6).number_format = "0.0"
    ws.cell(row=r, column=7).number_format = "0.0"
    ws.cell(row=r, column=8).number_format = "0.0%"
    ws.cell(row=r, column=2).number_format = "0"

widths = [13, 7, 12, 12, 13, 13, 15, 11]
for i, wd in enumerate(widths, start=1):
    ws.column_dimensions[chr(64 + i)].width = wd

ws["A13"] = "Blue = raw InBody readings. Fat Mass / Total Lost / % of Goal are derived (baseline = first weight, 154.7 kg)."
ws["A13"].font = Font(name=ARIAL, italic=True, size=9, color="808080")

# ---------- Injections ----------
wi = wb.create_sheet("Injections")
wi["A1"] = "Injection / Pin Log"
wi["A1"].font = Font(name=ARIAL, bold=True, size=14, color="FFFFFF")
wi["A1"].fill = TITLE_FILL
wi.merge_cells("A1:D1")
wi.row_dimensions[1].height = 24

ihead = ["Date", "Day", "Peptide", "Dose (mg)"]
for i, h in enumerate(ihead, start=1):
    c = wi.cell(row=3, column=i, value=h)
    c.font = Font(name=ARIAL, bold=True, color="FFFFFF")
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center")
    c.border = border

inj = [
    (date(2026, 4, 14), "Retatrutide", 1),
    (date(2026, 4, 17), "Retatrutide", 3),
    (date(2026, 5, 26), "Retatrutide", 4),
]
for idx, (d, p, mg) in enumerate(inj):
    r = 4 + idx
    wi.cell(row=r, column=1, value=d).number_format = "yyyy-mm-dd"
    wi.cell(row=r, column=2, value=(d - date(2026, 4, 14)).days)
    wi.cell(row=r, column=3, value=p)
    wi.cell(row=r, column=4, value=mg)
    for col in range(1, 5):
        cc = wi.cell(row=r, column=col)
        cc.border = border
        cc.alignment = Alignment(horizontal="center")
        cc.font = Font(name=ARIAL, color=(BLUE if col in (1, 3, 4) else BLACK))
    wi.cell(row=r, column=2).number_format = "0"
for i, wd in enumerate([13, 7, 16, 11], start=1):
    wi.column_dimensions[chr(64 + i)].width = wd

wb.save("/Users/thatsucks/shift/prometheus-seed-data.xlsx")
print("saved")
